import random
import sys
import string
from tkinter import *
from tkinter import messagebox
from PIL import Image

import customtkinter
from openpyxl import load_workbook

FILE = r'files\QMCounter.xlsx'

BLUE = "#d0e1f9"
RED = "#fe4a49"
GREEN = "#dcedc1"
YELLOW = "#fed766"
GRAY = "#eeeeee"
PURPLE = "#F0DEFD"

HL_BLUE = "#E6F0FD"
HL_RED = "#FA8F8F"
HL_GREEN = "#F2FAE6"
HL_YELLOW = "#FAE3A0"
HL_GRAY = "#FFFFFF"
HL_PURPLE = "#F6EEFB"

RED_2 = "#F60000"
GREEN_2 = "#00FF7F"
YELLOW_2 = "#f6cd61"
GRAY_2 = "#e3f0ff"

RED_3 = "#F7D8BA"
GREEN_3 = "#E1F8DC"
YELLOW_3 = "#FEF8DD"

FONT_COLOR = "#0000FF"
ALPHABET_LIST = list(string.ascii_uppercase)
STATUS_TYPES = {
    "O": GREEN,
    "N": RED,
    "Z": YELLOW,
    "T": BLUE,
    "U": PURPLE
}
STATUS_TYPES_HL = {
    "O": HL_GREEN,
    "N": HL_RED,
    "Z": HL_YELLOW,
    "T": HL_BLUE,
    "U": HL_PURPLE
}

main_window = None
ticket_counters = {}
employee_name_buttons = []
ticket_counter_buttons = []
employees = []
wb = None
ws = None


def get_employees():
    try:
        wt = load_workbook(FILE)
        wt.close()
    except FileNotFoundError:
        messagebox.showerror(title="Błąd", message="Plik arkusza z danymi nie istnieje lub ma zmienioną nazwę")
        sys.exit()
    else:
        global wb, ws
        wb = load_workbook(FILE)
        ws = wb.active
        try:
            wb.save(FILE)
        except PermissionError:
            messagebox.showerror(title="Błąd", message="Ktoś aktualnie ma otwarty plik arkusza z danymi")
            sys.exit()
        else:
            for i in range(2, 100):
                if ws['A' + str(i)].value:
                    employees.append(ws['A' + str(i)].value)
                    print(ws['A' + str(i)].value)
                else:
                    break
            wb.close()


def periodic_refresh():
    random_wait = random.randint(50, 70)
    main_window.after(random_wait * 1000, periodic_refresh)
    global wb, ws
    wb = load_workbook(FILE)
    ws = wb.active
    try:
        wb.save(FILE)
    except PermissionError:
        print("Ktoś aktualnie ma otwarty plik arkusza z danymi")
    else:
        num = 0
        for employee in employees:
            print(num)
            employee_name = employee
            employee_id = num + 2
            print(employee_name)
            for i in range(1, 4):
                current = ALPHABET_LIST[i] + str(employee_id)
                ticket_counters[employee_name]["entries"][i - 1].configure(state="normal")
                ticket_counters[employee_name]["entries"][i - 1].delete(0, END)
                ticket_counters[employee_name]["entries"][i - 1].insert(0, ws[current].value)
                ticket_counters[employee_name]["entries"][i - 1].configure(state="readonly")
            sum_up = int(ticket_counters[employee_name]["entries"][0].get()) \
                     + int(ticket_counters[employee_name]["entries"][1].get()) * 2 \
                     + int(ticket_counters[employee_name]["entries"][2].get()) * 3
            ticket_counters[employee_name]["entries"][3].configure(state="normal")
            ticket_counters[employee_name]["entries"][3].delete(0, END)
            ticket_counters[employee_name]["entries"][3].insert(0, sum_up)
            ticket_counters[employee_name]["entries"][3].configure(state="readonly")
            employee_name_buttons[num].configure(fg_color=STATUS_TYPES[ws["F" + str(employee_id)].value],
                                                 hover_color=STATUS_TYPES_HL[ws["F" + str(employee_id)].value])
            num += 1
    wb.close()


def refresh_data():
    try:
        wt = load_workbook(FILE)
        wt.close()
    except FileNotFoundError:
        messagebox.showerror(title="Błąd", message="Plik arkusza z danymi nie istnieje lub ma zmienioną nazwę")
        sys.exit()
    else:
        global wb, ws
        wb = load_workbook(FILE)
        ws = wb.active
        try:
            wb.save(FILE)
        except PermissionError:
            messagebox.showerror(title="Błąd", message="Ktoś aktualnie ma otwarty plik arkusza z danymi")
        else:
            num = 0
            for employee in employees:
                print(num)
                employee_name = employee
                employee_id = num + 2
                print(employee_name)
                for i in range(1, 4):
                    current = ALPHABET_LIST[i] + str(employee_id)
                    ticket_counters[employee_name]["entries"][i - 1].configure(state="normal")
                    ticket_counters[employee_name]["entries"][i - 1].delete(0, END)
                    ticket_counters[employee_name]["entries"][i - 1].insert(0, ws[current].value)
                    ticket_counters[employee_name]["entries"][i - 1].configure(state="readonly")
                sum_up = int(ticket_counters[employee_name]["entries"][0].get()) \
                         + int(ticket_counters[employee_name]["entries"][1].get()) * 2 \
                         + int(ticket_counters[employee_name]["entries"][2].get()) * 3
                ticket_counters[employee_name]["entries"][3].configure(state="normal")
                ticket_counters[employee_name]["entries"][3].delete(0, END)
                ticket_counters[employee_name]["entries"][3].insert(0, sum_up)
                ticket_counters[employee_name]["entries"][3].configure(state="readonly")
                employee_name_buttons[num].configure(fg_color=STATUS_TYPES[ws["F" + str(employee_id)].value],
                                                     hover_color=STATUS_TYPES_HL[ws["F" + str(employee_id)].value])
                num += 1
        wb.close()


def remove_data():
    try:
        wt = load_workbook(FILE)
        wt.close()
    except FileNotFoundError:
        messagebox.showerror(title="Błąd", message="Plik arkusza z danymi nie istnieje lub ma zmienioną nazwę")
        sys.exit()
    else:
        ask_box = messagebox.askquestion(title="Pytanie", message="Czy na pewno chcesz wyzerować wszystkie wartosci?")
        if ask_box == "yes":
            global wb, ws
            wb = load_workbook(FILE)
            ws = wb.active
            try:
                wb.save(FILE)
            except PermissionError:
                messagebox.showerror(title="Błąd", message="Ktoś aktualnie ma otwarty plik arkusza z danymi")
            else:
                num = 0
                for employee in employees:
                    print(num)
                    employee_name = employee
                    print(employee_name)
                    for i in range(1, 4):
                        ticket_counters[employee_name]["entries"][i - 1].configure(state="normal")
                        ticket_counters[employee_name]["entries"][i - 1].delete(0, END)
                        ticket_counters[employee_name]["entries"][i - 1].insert(0, 0)
                        ticket_counters[employee_name]["entries"][i - 1].configure(state="readonly")
                        ws[ALPHABET_LIST[i] + str(num + 2)] = 0
                    ticket_counters[employee_name]["entries"][3].configure(state="normal")
                    ticket_counters[employee_name]["entries"][3].delete(0, END)
                    ticket_counters[employee_name]["entries"][3].insert(0, 0)
                    ticket_counters[employee_name]["entries"][3].configure(state="readonly")
                    ws[ALPHABET_LIST[4] + str(num + 2)] = 0
                    num += 1
                wb.save(FILE)
            wb.close()


def add_count(num, num_e):
    try:
        wt = load_workbook(FILE)
        wt.close()
    except FileNotFoundError:
        messagebox.showerror(title="Błąd", message="Plik arkusza z danymi nie istnieje lub ma zmienioną nazwę")
        sys.exit()
    else:
        global wb, ws
        wb = load_workbook(FILE)
        ws = wb.active
        try:
            wb.save(FILE)
        except PermissionError:
            messagebox.showerror(title="Błąd", message="Ktoś aktualnie ma otwarty plik arkusza z danymi")
        else:
            print(num)
            employee_name = employees[num]
            employee_id = num + 2
            ticket_letter = ALPHABET_LIST[num_e + 1]
            print(employee_name)
            entry_value = ticket_counters[employee_name]["entries"][num_e].get()
            print(entry_value)
            entry_sum = int(entry_value) + 1
            ticket_counters[employee_name]["entries"][num_e].configure(state="normal")
            ticket_counters[employee_name]["entries"][num_e].delete(0, END)
            ticket_counters[employee_name]["entries"][num_e].insert(0, entry_sum)
            ticket_counters[employee_name]["entries"][num_e].configure(state="readonly")
            sum_up = int(ticket_counters[employee_name]["entries"][0].get()) \
                     + int(ticket_counters[employee_name]["entries"][1].get()) * 2 \
                     + int(ticket_counters[employee_name]["entries"][2].get()) * 3
            ticket_counters[employee_name]["entries"][3].configure(state="normal")
            ticket_counters[employee_name]["entries"][3].delete(0, END)
            ticket_counters[employee_name]["entries"][3].insert(0, sum_up)
            ticket_counters[employee_name]["entries"][3].configure(state="readonly")
            ws[ticket_letter + str(employee_id)] = entry_sum
            ws['E' + str(employee_id)] = sum_up
            wb.save(FILE)
        wb.close()


def rmv_count(num, num_e):
    try:
        wt = load_workbook(FILE)
        wt.close()
    except FileNotFoundError:
        messagebox.showerror(title="Błąd", message="Plik arkusza z danymi nie istnieje lub ma zmienioną nazwę")
        sys.exit()
    else:
        global wb, ws
        wb = load_workbook(FILE)
        ws = wb.active
        try:
            wb.save(FILE)
        except PermissionError:
            messagebox.showerror(title="Błąd", message="Ktoś aktualnie ma otwarty plik arkusza z danymi")
        else:
            print(num)
            employee_name = employees[num]
            employee_id = num + 2
            ticket_letter = ALPHABET_LIST[num_e + 1]
            print(employee_name)
            entry_value = ticket_counters[employee_name]["entries"][num_e].get()
            print(entry_value)
            if int(entry_value) > 0:
                entry_sum = int(entry_value) - 1
                ticket_counters[employee_name]["entries"][num_e].configure(state="normal")
                ticket_counters[employee_name]["entries"][num_e].delete(0, END)
                ticket_counters[employee_name]["entries"][num_e].insert(0, entry_sum)
                ticket_counters[employee_name]["entries"][num_e].configure(state="readonly")
                sum_up = int(ticket_counters[employee_name]["entries"][0].get()) \
                         + int(ticket_counters[employee_name]["entries"][1].get()) * 2 \
                         + int(ticket_counters[employee_name]["entries"][2].get()) * 3
                ticket_counters[employee_name]["entries"][3].configure(state="normal")
                ticket_counters[employee_name]["entries"][3].delete(0, END)
                ticket_counters[employee_name]["entries"][3].insert(0, sum_up)
                ticket_counters[employee_name]["entries"][3].configure(state="readonly")
                ws[ticket_letter + str(employee_id)] = entry_sum
                ws['E' + str(employee_id)] = sum_up
                wb.save(FILE)
        wb.close()


def update_status(num):
    try:
        wt = load_workbook(FILE)
        wt.close()
    except FileNotFoundError:
        messagebox.showerror(title="Błąd", message="Plik arkusza z danymi nie istnieje lub ma zmienioną nazwę")
        sys.exit()
    else:
        global wb, ws
        wb = load_workbook(FILE)
        ws = wb.active
        try:
            wb.save(FILE)
        except PermissionError:
            messagebox.showerror(title="Błąd", message="Ktoś aktualnie ma otwarty plik arkusza z danymi")
        else:
            current_color = employee_name_buttons[num].cget("fg_color")
            if current_color == BLUE:
                employee_name_buttons[num].configure(fg_color=RED, hover_color=HL_RED)
                ws["F" + str(num + 2)] = "N"
            elif current_color == RED:
                employee_name_buttons[num].configure(fg_color=GREEN, hover_color=HL_GREEN)
                ws["F" + str(num + 2)] = "O"
            elif current_color == GREEN:
                employee_name_buttons[num].configure(fg_color=YELLOW, hover_color=HL_YELLOW)
                ws["F" + str(num + 2)] = "Z"
            elif current_color == YELLOW:
                employee_name_buttons[num].configure(fg_color=PURPLE, hover_color=HL_PURPLE)
                ws["F" + str(num + 2)] = "U"
            else:
                employee_name_buttons[num].configure(fg_color=BLUE, hover_color=HL_BLUE)
                ws["F" + str(num + 2)] = "T"
            wb.save(FILE)
        wb.close()


def main_menu():
    global main_window, wb, ws
    customtkinter.set_appearance_mode("System")  # Modes: system (default), light, dark
    customtkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green
    main_window = customtkinter.CTk()
    main_window.title("QMCounter")
    main_window.config(padx=25, pady=25, bg=GRAY_2)
    main_window.resizable(False, False)

    logo_image = customtkinter.CTkImage(light_image=Image.open("files/logo.png"),
                                        dark_image=Image.open("files/logo.png"),
                                        size=(250, 75))

    image_label = customtkinter.CTkLabel(main_window, image=logo_image, text="", bg_color=GRAY_2)  # display image
    # with a CTkLabel
    image_label.grid(pady=5, padx=5, row=0, column=0)

    last_row = None

    # Employee Names

    place = 1
    btn_place = 1
    for i, employee in enumerate(employees, start=1):
        employee_name_button = customtkinter.CTkButton(main_window, text=employee, font=("Times", 24, "normal"),
                                                       width=250, border_width=2, height=75, text_color="black",
                                                       fg_color=GREEN, bg_color=GRAY_2,
                                                       command=lambda employee_id=place - 1: update_status(employee_id))
        employee_name_button.grid(pady=10, row=btn_place, column=0, rowspan=2)
        calculated_value = ws['B' + str(place + 1)].value + ws['C' + str(place + 1)].value * 2 + ws[
            'D' + str(place + 1)].value * 3
        ticket_counters[employee] = {}
        ticket_counters[employee]["tickets"] = [ws['B' + str(place + 1)].value, ws['C' + str(place + 1)].value,
                                                ws['D' + str(place + 1)].value, calculated_value]
        ticket_counters[employee]["buttons-add"] = [None, None, None, None]
        ticket_counters[employee]["buttons-rmv"] = [None, None, None, None]
        ticket_counters[employee]["entries"] = [None, None, None, None]
        ticket_counters[employee]["status"] = ws["F" + str(place + 1)]
        employee_name_button.configure(fg_color=STATUS_TYPES[ws["F" + str(place + 1)].value],
                                       hover_color=STATUS_TYPES_HL[ws["F" + str(place + 1)].value])
        employee_name_buttons.append(employee_name_button)
        place += 1
        btn_place += 2
        last_row = btn_place

    # Number Headers

    place = 1
    for i in range(1, 6, 2):
        number_header = customtkinter.CTkButton(main_window, text=str(place), font=("Times", 30, "normal"),
                                                fg_color="white", bg_color=GRAY_2,
                                                state="disabled", border_width=2, height=100, width=100,
                                                text_color_disabled="black")
        if i == 1:
            number_header.configure(fg_color=GREEN_3)
        elif i == 3:
            number_header.configure(fg_color=YELLOW_3)
        else:
            number_header.configure(fg_color=RED_3)
        number_header.grid(pady=1, row=0, column=i, columnspan=2)
        place += 1
    sum_header = customtkinter.CTkButton(main_window, text="=", font=("Times", 30, "normal"), fg_color="white",
                                         state="disabled", border_width=2, height=100, width=100, bg_color=GRAY_2,
                                         text_color_disabled="black")
    sum_header.grid(pady=1, row=0, column=7, columnspan=2)

    place = 1
    btn_place = 1
    add_place = 1
    rmv_place = 2
    for employee in employees:
        number = 0
        for i in range(1, 6, 2):
            ticket_counter = customtkinter.CTkEntry(main_window, font=("Times", 36, "normal"), width=50, height=70,
                                                    justify="center",
                                                    state="normal", text_color="black", fg_color="white",
                                                    exportselection=False, border_width=1, bg_color=GRAY_2)
            ticket_counter.insert(0, ticket_counters[employee]["tickets"][number])
            if i == 1:
                ticket_counter.configure(fg_color=GREEN_3)
            elif i == 3:
                ticket_counter.configure(fg_color=YELLOW_3)
            else:
                ticket_counter.configure(fg_color=RED_3)
            ticket_counter.configure(state="readonly")
            ticket_counter.grid(padx=1, row=btn_place, column=i, rowspan=2)
            ticket_counters[employee]["entries"][number] = ticket_counter
            # Add tickets (+1)
            add_counter = customtkinter.CTkButton(main_window, text="+", font=("Times", 24, "normal"),
                                                  text_color="black", fg_color=GRAY, hover_color=HL_GREEN,
                                                  border_width=2, width=60, height=30, corner_radius=25,
                                                  bg_color=GRAY_2,
                                                  command=lambda employee_id=place - 1, entry_id=number: add_count(
                                                      employee_id,
                                                      entry_id))
            add_counter.grid(padx=1, row=add_place, column=i + 1)
            ticket_counters[employee]["buttons-add"][number] = add_counter
            ticket_counter_buttons.append(ticket_counter)
            # Remove tickets (-1)
            rmv_counter = customtkinter.CTkButton(main_window, text="-", font=("Times", 24, "normal"),
                                                  text_color="black", fg_color=GRAY, hover_color=HL_RED,
                                                  bg_color=GRAY_2,
                                                  border_width=2, width=60, height=30, corner_radius=25,
                                                  command=lambda employee_id=place - 1, entry_id=number: rmv_count(
                                                      employee_id,
                                                      entry_id))
            rmv_counter.grid(padx=1, row=rmv_place, column=i + 1)
            ticket_counters[employee]["buttons-rmv"][number] = rmv_counter
            ticket_counter_buttons.append(ticket_counter)

            number += 1
        ticket_counter = customtkinter.CTkEntry(main_window, font=("Times", 36, "normal"), width=80, height=70,
                                                justify="center",
                                                state="normal", text_color="black", fg_color="white",
                                                exportselection=False, border_width=1, bg_color=GRAY_2)
        ticket_counter.insert(0, ticket_counters[employee]["tickets"][number])
        ticket_counter.configure(state="readonly")
        ticket_counter.grid(padx=1, row=btn_place, column=7, rowspan=2)
        ticket_counters[employee]["entries"][number] = ticket_counter
        place += 1
        btn_place += 2
        add_place += 2
        rmv_place += 2

    refresh_button = customtkinter.CTkButton(main_window, text="Pobierz aktualne dane", font=("Times", 24, "normal"),
                                             width=720, height=50, bg_color=GRAY_2,
                                             fg_color=GRAY, text_color="black", hover_color="white", border_width=1,
                                             command=refresh_data)
    refresh_button.grid(pady=5, padx=5, row=last_row, column=0, columnspan=8)
    remove_button = customtkinter.CTkButton(main_window, text="Wyzeruj wszystkie dane", font=("Times", 24, "normal"),
                                            width=720, height=50, bg_color=GRAY_2,
                                            fg_color=GRAY, text_color="black", hover_color="white", border_width=1,
                                            command=remove_data)
    remove_button.grid(pady=5, padx=5, row=last_row + 1, column=0, columnspan=8)

    periodic_refresh()
    main_window.mainloop()


get_employees()
main_menu()
wb.close()
